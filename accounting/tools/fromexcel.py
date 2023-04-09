from openpyxl import load_workbook

def from_excel(filename, classes_by_name):
    '''создает иерархию классов на основании данных из excel'''
    wb = load_workbook(filename)
    # создаем словарь temp_data с данными считанными из excel,
    # в нем по ключам - названиям классов (названия листов), 
    # хранятся списки - данные для свойств экземпляров 
    # (заголовки,  строки)
    temp_data = load_temp_data(wb)
    instances_by_id = create_instances(temp_data, classes_by_name)
    create_relationships(temp_data, instances_by_id)


def load_temp_data(wb):
    '''создает временную структуру для хранения данных из excel'''
    temp_data = {}
    # проходим по всем листам книги
    for ws in wb:
        # название листа является названием класса экземпляров
        class_name = ws.title
        # считываем названия столбцов листа
        headers = [cell.value for cell in ws[1]]
        
        temp_data[class_name] = []
        # проходим по всем строкам, исключая заголовки
        for row in ws.iter_rows(min_row=2):
            # получаем данные для экземпляра в виде словаря 
            # {'имя свойства', значение}
            instance_data = get_instance_data(row, headers)
            # добавляем эти данные в список свойств для данного класса
            temp_data[class_name].append(instance_data)
    return temp_data


def get_instance_data(row, headers):
    '''получает значения для свойств экземпляра по заголовкам столбца 
    и строки данных'''
    instance_data = {}
    # zip объединеняет списки в список кортежей, попарно
    # т.е. заданных заголовков и данных получаем пары - 
    # (название заголовка, значение)
    for header, cell in zip(headers, row):
            # в словаре по названию заголовка сохраняем значение 
            instance_data[header] = cell.value
    return instance_data


def create_instances(temp_data, classes_by_name):
    '''создает экземпляры со свойствами из field, но без дочерних элементов,
    сохраняет экземпляры в словаре на кортежах (имя класса, номер экземпляра)'''
    instances_by_id = {}
    # проходим по всем классам по парам (имя класса, список наборов данных для экземпляров)
    for class_name, instances_data in temp_data.items():
        # classes_by_name словарь в котором хранятся актуальные типы классов
        # по их названиям, заполняется в main.py, т.е. здесь получаем тип класса
        instance_class = classes_by_name[class_name]
        # проходим по списку наборов данных для экземпляров текущего класса
        for properties_data in instances_data:
            # получаем номер экземпляра в текущем наборе
            instance_id = properties_data["id_by_class"]
            # если пара (название класса, номер экземпляра) уникальна, то
            if (class_name, instance_id) not in instances_by_id:
                # создаем экземпляр текущего класса с текущим порядковым номером
                instance = instance_class(instance_id)
                # проходим по всем элементам списка свойств экземпляра
                for property_name, property_value in properties_data.items():
                    # если это простые данные (нет постфикса), то записываем эту  
                    # пару в свойство экземпляра field
                    if not property_name.endswith("_instance"):
                        instance.fields[property_name]=property_value

                    # сохраняем экземпляр в словаре на кортеже, дочерние
                    # классы subinstances заполнятся позже в create_relationships
                    instances_by_id[(class_name, instance_id)] = instance
    return instances_by_id


def create_relationships(temp_data, instances_by_id):
    ''' создает связи между экземплярами, добавляя в subinstances объекты дочерних экземпляров 
    temp_data - список наборов свойств экземпляров [{'ClassName1':{'PropertyName':PropertyValue, ...}}, ...]; 
    instances_by_id - словарь на списках кортежей ('ClassName1', 'id_by_class') = <class ClassName1 object>'''
    # проходим по всем классам по парам (имя класса, список наборов данных для экземпляров)
    for class_name, instances_data in temp_data.items():
            # проходим по списку наборов данных для экземпляров текущего класса
            for properties_data in instances_data:
                # каждый набор данных это экземпляр с порядковым номером
                instance_id = properties_data["id_by_class"]
                # получаем объект данного экземпляра из словаря по кортежу
                instance = instances_by_id[(class_name, instance_id)]
                # проходим по всем элементам списка свойств экземпляра
                for property_name, property_value in properties_data.items():

                    # если есть постфикс, то элемент является экземпляром класса
                    # значит его нужно добавить в свойство subinstances
                    if property_name.endswith("_instance"):
                        # убираем постфикс из его названия
                        property_name = property_name[:-len("_instance")]
                        
                        # если указано несколько номеров экземпляров через запятую,
                        # запятые пока не используются, на будущее
                        subinstance_ids = str(property_value).split(',')  
                        # создаем экземпляры для всех порядковых номеров
                        for subinstance_id in subinstance_ids:
                            if subinstance_id != 'None':
                                subinstance = instances_by_id[(property_name, int(subinstance_id))]
                                instance.add_subinstance(subinstance)
    pass