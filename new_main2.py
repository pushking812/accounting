from openpyxl import Workbook, load_workbook

def get_headers(instances):
    headers = set()
    for instance in instances:
        headers.update(instance.fields.keys())
        for subinstances_class_name in instance.subinstances.keys():
            headers.add(f"{subinstances_class_name}_instance")
    return list(headers)

def get_row(instance, headers):
    row = []
    for header in headers:
        if header.endswith("_instance"):
            subinstances_class_name = header[:-len("_instance")]
            subinstances = instance.subinstances.get(subinstances_class_name, [])
            row.append(",".join(str(subinstance.fields["id_by_class"]) for subinstance in subinstances))
        else:
            row.append(instance.fields.get(header))
    return row

def get_instance_data(row, headers):
    instance_data = {}
    for header, cell in zip(headers, row):
        # if header.endswith("_instance"):
        #     subinstances_class_name = header[:-len("_instance")]
        #     subinstances_ids = cell.value.split(",") if cell.value else []
        #     instance_data[subinstances_class_name] = subinstances_ids
        # else:
            instance_data[header] = cell.value
    return instance_data

def load_temp_data(wb):
    temp_data = {}
    for ws in wb:
        class_name = ws.title
        headers = [cell.value for cell in ws[1]]
        temp_data[class_name] = []
        for row in ws.iter_rows(min_row=2):
            instance_data = get_instance_data(row, headers)
            temp_data[class_name].append(instance_data)
    # print(temp_data)  # Print out the temp_data dictionary for debugging
    return temp_data


def create_instances(temp_data):
    instances_by_id = {}
    for class_name, instances_data in temp_data.items():
        instance_class = globals()[class_name]
        for instance_data in instances_data:
            instance_id = instance_data["id_by_class"]
            if (class_name, instance_id) not in instances_by_id:
                instance = instance_class(instance_id)
                for key, value in instance_data.items():
                    if not key.endswith("_instance"):
                        instance.fields[key]=value
                instances_by_id[(class_name, instance_id)] = instance
    return instances_by_id


def create_relationships(temp_data, instances_by_id):
    for class_name, instances_data in temp_data.items():
            for instance_data in instances_data:
                instance_id = instance_data["id_by_class"]
                instance = instances_by_id[(class_name, instance_id)]
                for subinstances_class_name, subinstances_ids in instance_data.items():
                    if subinstances_class_name.endswith("_instance"):
                        subinstances_class_name = subinstances_class_name[:-len("_instance")]
                        for subinstance_id in subinstances_ids:
                            subinstance = instances_by_id[(subinstances_class_name, int(subinstance_id))]
                            # print(subinstances_class_name, subinstances_ids, subinstance)
                            instance.add_subinstance(subinstance)
    pass

class CountInstances:
    count_by_class = {}
    max_id_by_class = {}

    @classmethod
    def add_instance(cls, instance):
        class_name = instance.__class__.__name__
        cls.count_by_class[class_name] = cls.count_by_class.get(class_name, 0) + 1
        cls.max_id_by_class[class_name] = cls.max_id_by_class.get(class_name, 0) + 1
        return cls.max_id_by_class[class_name]
    
    @classmethod
    def remove_instance(cls, instance):
        class_name = instance.__class__.__name__
        if class_name in cls.count_by_class:
            cls.count_by_class[class_name] -= 1
        
    @classmethod
    def get_number(cls, instance):
        class_name = instance.__class__.__name__
        return cls.count_by_class.get(class_name, 0)

class Instance:
    instances_by_class = {}

    def __init__(self, instance_class, instance_id = None):
        class_name = instance_class.__name__

        instances = Instance.instances_by_class
        self.fields = {}
        
        if class_name not in instances:
            instances[class_name] = []

        instances[class_name].append(self)

        if instance_id == None:
            instance_id=CountInstances.add_instance(self)

        self.fields = {
            'id_by_class': instance_id
        }

        self.subinstances = {}
    
    # def add_subinstance(self, instance):
    #     key = type(instance).__name__
    #     if key not in self.subinstances:
    #         self.subinstances[key] = []
    #     self.subinstances[key].append(instance)

    def add_subinstance(self, instance):
        key = type(instance).__name__
        if key not in self.subinstances:
            self.subinstances[key] = []
        if instance not in self.subinstances[key]:
            self.subinstances[key].append(instance)
            # pass

    def get_subinstance(self, key, index=None):
        if key in self.subinstances:
            value = self.subinstances[key]
            if isinstance(value, list):
                if index is None:
                    raise ValueError("Index hasn't been set")
                return value[index]
            return value
        return []

    def remove_subinstance(self, instance):
        key = type(instance).__name__
        if key in self.subinstances:
            self.subinstances[key].remove(instance)
            Instance.instances_by_class[key].remove(instance)
            CountInstances.remove_instance(instance)

    @classmethod
    def display_properties(cls, instance):
        for n in instance.__dict__.items():
            print('----------------------')
            print(n)
        print('========================')

    @classmethod
    def to_excel(cls, filename):
        wb = Workbook()
        for class_name, instances in cls.instances_by_class.items():
            ws = wb.create_sheet(class_name)
            headers = get_headers(instances)
            ws.append(headers)
            for instance in instances:
                row = get_row(instance, headers)
                ws.append(row)
        wb.save(filename)
    
    @classmethod
    def from_excel(cls, filename):
        wb = load_workbook(filename)
        temp_data = load_temp_data(wb)
        instances_by_id = create_instances(temp_data)
        create_relationships(temp_data, instances_by_id)

class Record(Instance):
    def __init__(self, cls, instance_id = None, subinstance = None, commentary = None):
        super().__init__(cls, instance_id)

        if commentary == None:
            commentary = ''
        
        self.fields['commentary'] = commentary

        if subinstance != None:
            self.add_subinstance(subinstance)

class ObjectRecord(Record):
    def __init__(self, instance_id = None, subinstance = None, commentary = None):
        super().__init__(type(self), instance_id, subinstance, commentary)
 
class ProjectRecord(Record):
    def __init__(self, instance_id = None, subinstance = None, commentary = None):
        super().__init__(type(self), instance_id, subinstance, commentary)
      
class Object(Instance):
    def __init__(self, instance_id = None, object_number = None, object_name = None):
        super().__init__(type(self), instance_id)
        
        if object_number == None:
            object_number = ''

        if object_name == None:
            object_name = ''

        self.fields['object_number'] = object_number
        self.fields['object_name'] = object_name

class Project(Instance):
    def __init__(self, instance_id = None, project_number = None, project_name = None):
        super().__init__(type(self), instance_id)

        if project_number == None:
            project_number = ''

        if project_name == None:
            project_name = ''

        self.fields['project_number'] = project_number
        self.fields['project_name'] = project_name

def to_new_file():
    # заполняем справочники
    o = Object(None, 'O1', 'Object1')
    p1 = Project(None, 'P1', 'Project1')
    p2 = Project(None, 'P2', 'Project2')
    p3 = Project(None, 'P3', 'Project3')
    # заполяем записи данными
    obr1 = ObjectRecord(None, o)
    obr2 = ObjectRecord(None, o)
    obr3 = ObjectRecord(None, o)
    prr1 = ProjectRecord(None, p1)
    prr2 = ProjectRecord(None, p2)
    prr3 = ProjectRecord(None, p3)
    # организуем связи между данными
    obr1.add_subinstance(prr1)
    obr2.add_subinstance(prr2)
    obr3.add_subinstance(prr3)
    Object.to_excel('data.xlsx')
    Project.to_excel('data.xlsx')
    ObjectRecord.to_excel('data.xlsx')
    ProjectRecord.to_excel('data.xlsx')

def read_from_file():
    Instance.from_excel('data.xlsx')
    # Project.from_excel('data.xlsx')
    # ObjectRecord.from_excel('data.xlsx')
    # ProjectRecord.from_excel('data.xlsx')
    # objrec = Instance.instances_by_class['ObjectRecord']
    # for obj in objrec:
    #     print(obj.fields)
    #     print(obj.subinstances)

def write_to_file():
    Object.to_excel('data2.xlsx')
    # Project.to_excel('data2.xlsx')
    # ObjectRecord.to_excel('data2.xlsx')
    # ProjectRecord.to_excel('data2.xlsx')

# to_new_file()

read_from_file()
write_to_file()

