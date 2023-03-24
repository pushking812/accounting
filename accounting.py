import tkinter as tk
from tkinter import messagebox

from zipfile import BadZipFile

import sys
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import json
import os

# Загрузка списка имен файлов спецификаций из конфигурационного файла
def load_spec_files():
    try:
        with open("config.json") as f:
            data = json.load(f)
            return data["spec_files"]
    except FileNotFoundError:
        fatal_error(FileNotFoundError, "config.json file not found")
    except json.JSONDecodeError:
        fatal_error(json.JSONDecodeError, "Invalid JSON format in config.json")
    except KeyError:
        fatal_error(KeyError, "'spec_files' key not found in config.json")
    print("config.json loaded!")

def fatal_error(err, msg):
        messagebox.showerror("Error", f"{err}\n{msg}")
        print(f"Error: {err}\n{msg}")
        sys.exit(1)

# Проверка существования файла и корректности его формата (Excel)
def check_files(file_list):
    for file_name in file_list:
        if not os.path.exists(file_name):
            fatal_error("", f"Could not open file {file_name}")
        else:
            try:
                openpyxl.load_workbook(file_name)
            except (InvalidFileException, BadZipFile) as err:
                fatal_error(err, f"{file_name} is not a valid Excel file")
    print("All files have been checked!")

# Загрузка данных из спецификаций в справочник
def load_spec_dictionary(spec_files):
    spec_dict = {}
    empty_value = False
    for file_name in spec_files:
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    item_number = row[1]
                    volume = row[3]
                    unit = row[4]

                    if item_number == None or volume == None or unit == None:
                        empty_value = True
                        print(f"Error: Cell values of row {row} in file {file_name} are empty")

                    spec_dict[item_number] = (file_name, volume, unit)
                except IndexError:
                    fatal_error(IndexError, f"Error: Row {row} in file {file_name} is missing data")
        except InvalidFileException:
            fatal_error(InvalidFileException, f"Could not open file {file_name}")
  
    if empty_value == True:
        fatal_error("", f"Error: Rows of file {file_name} have empty values")

    return spec_dict

# Отладочная печать справочника
def print_dict(dict):
    for item_number, (file_name, volume, unit) in dict.items():
        print(f"File name: {file_name}, Item number: {item_number}, Volume: {volume}, Unit: {unit}")

invoice_dict = {}

# Загрузка данных из накладной в справочник 
def load_invoice(file_name):
    global invoice_dict
    try:
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                item_number = row[1]
                volume = row[3]
                unit = row[4]
                invoice_dict[item_number] = (file_name, volume, unit)
            except IndexError:
                messagebox.showerror("Error", f"Error: {IndexError}\nRow {row} is missing data")
                print(f"Error: {IndexError}:\nRow {row} is missing data")
                return
    except InvalidFileException:
        messagebox.showerror("Error", f"Error: {InvalidFileException}\nCould not open file {file_name}")
        print(f"Error: {InvalidFileException}\nCould not open file {file_name}")
        return
    messagebox.showinfo("Information", f"Invoice {file_name} loaded")
    print(f"Invoice {file_name} loaded")
    print_dict(invoice_dict)


def merge_dicts_by_key(dict1, dict2):
    merged_dict = {}
    not_found_dict = {}
    for key, value in dict2.items():
        if key in dict1:
            merged_dict[key] = {**dict1[key], **value}
        else:
            not_found_dict[key] = value
    return merged_dict, not_found_dict

# Активирование требуемого фрейма
def raise_page(page):
    page.tkraise()

# Загрузка и проверка данных
spec_files = load_spec_files()
check_files(spec_files)
spec_dict = load_spec_dictionary(spec_files)
print_dict(spec_dict)


# GUI интерфейс многостраничного диалога

root = tk.Tk()
root.title("Учет материалов и оборудования")

# Фреймы многостраничного диалога
page1 = tk.Frame(root)
page2 = tk.Frame(root)
page3 = tk.Frame(root)

# Виджеты фрейма 1
invoice_label = tk.Label(page1, text="Файл накладной:")
invoice_entry = tk.Entry(page1)
load_button = tk.Button(page1, text="Загрузить", command=lambda: load_invoice(invoice_entry.get()))
next_button = tk.Button(page1, text="Далее", command=lambda: raise_page(page2))

# Размещение виджетов фрейма 1
invoice_label.grid(row=0, column=0)
invoice_entry.grid(row=0, column=1)
load_button.grid(row=1, column=0)
next_button.grid(row=2, column=0)

# Виджеты фрейма 2
match_label = tk.Label(page2, text="Сопоставить элементы")
back_button = tk.Button(page2, text="Назад", command=lambda: raise_page(page1))
next_button2 = tk.Button(page2, text="Далее", command=lambda: raise_page(page3))

# Размещение виджетов фрейма 2
match_label.grid(row=0, column=0)
back_button.grid(row=1, column=0)
next_button2.grid(row=1, column=1)

# Виджеты фрейма 3
review_label = tk.Label(page3, text="Просмотр сопоставлений")
back_button2 = tk.Button(page3, text="Назад", command=lambda: raise_page(page2))

# Размещение виджетов фрейма 3
review_label.grid(row=0, column=0)
back_button2.grid(row=1, column=0)

# Сложите страницы друг на друга
for page in (page1, page2, page3):
    page.grid(row=0, column=0, sticky="nsew")

# Поднимите страницу 1 на верх
raise_page(page1)

root.mainloop()
